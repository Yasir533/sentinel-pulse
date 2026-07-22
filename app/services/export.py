import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ReportLab Imports
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

class ExportService:
    @staticmethod
    def export_csv(headers: list, rows: list) -> str:
        """
        Export a list of records to CSV format.
        """
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        csv_str = output.getvalue()
        output.close()
        return csv_str

    @staticmethod
    def export_xlsx(sheet_name: str, headers: list, rows: list) -> bytes:
        """
        Export a list of records to Excel format.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:30] # Excel sheet title limits
        ws.views.sheetView[0].showGridLines = True

        title_font = Font(name="Calibri", size=14, bold=True, color="0EA5E9")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Calibri", size=11)
        
        header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # Title block
        ws["A1"] = f"Sentinel Pulse - {sheet_name} Export"
        ws["A1"].font = title_font
        ws["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        ws["A2"].font = Font(name="Calibri", size=9, italic=True)

        # Write Headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Write Rows
        for row_idx, row_data in enumerate(rows, 5):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Normalize values
                if isinstance(val, datetime):
                    cell.value = val.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    cell.value = str(val) if val is not None else ""
                
                cell.font = regular_font
                cell.border = thin_border
                if row_idx % 2 == 1:
                    cell.fill = zebra_fill

        # Auto-adjust column width
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        out = io.BytesIO()
        wb.save(out)
        xlsx_bytes = out.getvalue()
        out.close()
        return xlsx_bytes

    @staticmethod
    def export_pdf(title: str, headers: list, rows: list) -> bytes:
        """
        Export a list of records to PDF format using landscape letter size.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )

        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            name='ExportTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=18,
            textColor=colors.HexColor('#0ea5e9'),
            spaceAfter=5
        )
        
        meta_style = ParagraphStyle(
            name='ExportMeta',
            fontName='Helvetica-Oblique',
            fontSize=8,
            textColor=colors.HexColor('#64748b'),
            spaceAfter=15
        )

        header_th_style = ParagraphStyle(
            name='THStyle',
            fontName='Helvetica-Bold',
            fontSize=8,
            textColor=colors.white,
            alignment=1  # Centered
        )

        cell_style = ParagraphStyle(
            name='TDStyle',
            fontName='Helvetica',
            fontSize=8,
            textColor=colors.HexColor('#0f172a'),
            leading=10
        )

        story = []
        story.append(Paragraph("SENTINEL PULSE REGISTRY", ParagraphStyle('Sub', fontName='Helvetica-Bold', fontSize=8, textColor=colors.HexColor('#64748b'))))
        story.append(Paragraph(title.upper(), title_style))
        story.append(Paragraph(f"Export Date: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}", meta_style))

        # Format Table data (convert strings to Paragraphs for word wrapping)
        table_data = []
        
        # Header Row
        header_row = [Paragraph(h, header_th_style) for h in headers]
        table_data.append(header_row)
        
        # Data Rows
        for r in rows:
            formatted_row = []
            for item in r:
                val_str = item.strftime('%Y-%m-%d %H:%M') if isinstance(item, datetime) else str(item) if item is not None else ""
                # Escape HTML special chars just in case
                val_str = val_str.replace('<', '&lt;').replace('>', '&gt;')
                formatted_row.append(Paragraph(val_str, cell_style))
            table_data.append(formatted_row)

        # Dynamic col widths based on landscape space (approx 730 points printable width)
        col_count = len(headers)
        col_width = int(730 / col_count)
        widths = [col_width] * col_count

        records_table = Table(table_data, colWidths=widths, repeatRows=1)
        records_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0ea5e9')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
            ('PADDING', (0,0), (-1,-1), 6),
        ]))
        
        story.append(records_table)
        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes

    @classmethod
    def export_to_csv(cls, entity_type: str, filters: dict) -> bytes:
        """Fetch entity data and export to CSV format bytes."""
        headers, rows = cls._get_entity_data(entity_type, filters)
        csv_str = cls.export_csv(headers, rows)
        return csv_str.encode('utf-8')

    @classmethod
    def export_to_xlsx(cls, entity_type: str, filters: dict) -> bytes:
        """Fetch entity data and export to Excel format bytes."""
        headers, rows = cls._get_entity_data(entity_type, filters)
        return cls.export_xlsx(entity_type.capitalize(), headers, rows)

    @classmethod
    def export_to_pdf(cls, entity_type: str, filters: dict) -> bytes:
        """Fetch entity data and export to PDF format bytes."""
        headers, rows = cls._get_entity_data(entity_type, filters)
        return cls.export_pdf(entity_type.capitalize(), headers, rows)

    @staticmethod
    def _get_entity_data(entity_type: str, filters: dict) -> tuple[list, list]:
        """Dynamically fetch columns and rows for the specified entity type for exporting."""
        from app.models.threat import Threat
        from app.models.alert import Alert
        from app.models.incident import Incident
        from app.models.user import User
        from app.models.audit_log import AuditLog
        from app.models.notification import Notification

        if entity_type == 'threats':
            records = Threat.query.all()
            headers = ["Ingestion Date", "Threat Type", "IOC Type", "Indicator Value", "Severity", "Status", "Source Feed"]
            rows = [[t.created_at, t.threat_type, t.ioc_type, t.ioc_value, t.severity, t.status, t.source] for t in records]
        elif entity_type == 'alerts':
            records = Alert.query.all()
            headers = ["Alert Number", "Triggered Timestamp", "Message", "Severity", "Status"]
            rows = [[a.alert_number, a.created_at, a.message, a.severity, a.status] for a in records]
        elif entity_type == 'incidents':
            records = Incident.query.all()
            headers = ["Incident Title", "Created At", "Description", "Severity", "Status"]
            rows = [[i.title, i.created_at, i.description, i.severity, i.status] for i in records]
        elif entity_type == 'users':
            records = User.query.all()
            headers = ["Username", "Email Address", "Operator Role", "Is Active", "Registered At"]
            rows = [[u.username, u.email, u.role, u.is_active, u.created_at] for u in records]
        elif entity_type == 'notifications':
            records = Notification.query.all()
            headers = ["Timestamp", "Notification message", "Priority", "Status"]
            rows = [[n.created_at, n.message, n.priority, n.status] for n in records]
        elif entity_type == 'audit_logs':
            records = AuditLog.query.all()
            headers = ["Audit Date & Time", "Username", "Operator Role", "Remote IP", "Action Executed", "Target Entity", "Execution status"]
            rows = [[log.timestamp, log.username, log.role, log.ip_address, log.action, log.entity, log.status] for log in records]
        else:
            headers, rows = [], []
        return headers, rows

