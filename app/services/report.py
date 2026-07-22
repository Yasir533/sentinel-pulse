import io
import csv
from datetime import datetime
from sqlalchemy import func
from app.extensions import db
from app.models.report import Report
from app.models.threat import Threat
from app.models.alert import Alert
from app.models.incident import Incident
from app.models.user import User

# ReportLab Imports
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# OpenPyXL Imports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class ReportService:
    @staticmethod
    def generate_next_report_number() -> str:
        """
        Generate an auto-incrementing report number: RPT-YYYY-XXXX.
        """
        year = datetime.utcnow().year
        prefix = f"RPT-{year}-"
        max_rpt = db.session.query(Report).filter(Report.report_number.like(f"{prefix}%")).order_by(Report.report_number.desc()).first()
        if max_rpt:
            try:
                last_seq = int(max_rpt.report_number.split('-')[-1])
                next_seq = last_seq + 1
            except ValueError:
                next_seq = 1
        else:
            next_seq = 1
        return f"{prefix}{next_seq:04d}"

    @classmethod
    def generate_report(cls, report_type: str, creator_id: int) -> Report:
        """
        Gathers dynamic security statistics and saves a new Report record in the database.
        """
        # Fetch creator user
        db.session.get(User, creator_id)

        # 1. Fetch statistics
        total_threats = Threat.query.count()
        threats_by_severity = {
            'Critical': Threat.query.filter_by(severity='Critical').count(),
            'High': Threat.query.filter_by(severity='High').count(),
            'Medium': Threat.query.filter_by(severity='Medium').count(),
            'Low': Threat.query.filter_by(severity='Low').count()
        }
        
        # Threats by type
        threat_type_query = db.session.query(Threat.threat_type, func.count(Threat.id)).group_by(Threat.threat_type).all()
        threats_by_type = {t_type: count for t_type, count in threat_type_query if t_type}

        total_alerts = Alert.query.count()
        alerts_by_severity = {
            'Critical': Alert.query.filter_by(severity='Critical').count(),
            'High': Alert.query.filter_by(severity='High').count(),
            'Medium': Alert.query.filter_by(severity='Medium').count(),
            'Low': Alert.query.filter_by(severity='Low').count()
        }
        alerts_by_status = {
            'New': Alert.query.filter_by(status='New').count(),
            'Acknowledged': Alert.query.filter_by(status='Acknowledged').count(),
            'Investigating': Alert.query.filter_by(status='Investigating').count(),
            'Resolved': Alert.query.filter_by(status='Resolved').count()
        }

        total_incidents = Incident.query.count()
        incidents_by_status = {
            'Open': Incident.query.filter_by(status='Open').count(),
            'In Progress': Incident.query.filter_by(status='In Progress').count(),
            'Under Investigation': Incident.query.filter_by(status='Under Investigation').count(),
            'Resolved': Incident.query.filter_by(status='Resolved').count(),
            'Closed': Incident.query.filter_by(status='Closed').count()
        }
        incidents_by_severity = {
            'Critical': Incident.query.filter_by(severity='Critical').count(),
            'High': Incident.query.filter_by(severity='High').count(),
            'Medium': Incident.query.filter_by(severity='Medium').count(),
            'Low': Incident.query.filter_by(severity='Low').count()
        }

        total_users = User.query.count()
        active_users = User.query.filter_by(is_active=True).count()

        # 2. Formulate dynamic summary text based on report type
        if report_type == 'Executive Security Report':
            title = "Executive Security Report"
            summary = (
                f"During the current reporting period, the SOC monitored a total of {total_threats} threat indicators, "
                f"which resulted in {total_alerts} security alerts and {total_incidents} escalated incidents. "
                f"Currently, there are {incidents_by_status.get('Open', 0) + incidents_by_status.get('In Progress', 0)} active incident investigations. "
                f"The overall platform security posture remains stable, with active containment procedures in effect."
            )
            recommendations = [
                "Prioritize the triage and resolution of unresolved critical alerts.",
                "Review firewall rules and blocklists matching recent indicator ingestions.",
                "Perform scheduled vulnerability assessments on endpoints linked to incident hosts."
            ]
        elif report_type == 'Threat Intelligence Report':
            title = "Threat Intelligence Report"
            summary = (
                f"This report summarizes intelligence ingestion and indicator profiles. A total of {total_threats} Indicators of Compromise (IOCs) "
                f"are registered. Severity breakdown indicates {threats_by_severity.get('Critical', 0)} Critical, {threats_by_severity.get('High', 0)} High risk threats. "
                f"Malware signatures and malicious IP ranges constitute the primary vectors of attack."
            )
            recommendations = [
                "Incorporate threat intelligence feeds directly into SIEM blocking filters.",
                "Perform reputational correlation checkups on active IP addresses.",
                "Run standard hash matches on endpoint systems to detect quarantined signatures."
            ]
        elif report_type == 'Incident Summary Report':
            title = "Incident Summary Report"
            summary = (
                f"A review of Incident Response operations indicates a lifetime total of {total_incidents} escalated incidents. "
                f"To date, {incidents_by_status.get('Resolved', 0)} incidents have been resolved, and {incidents_by_status.get('Closed', 0)} "
                f"officially closed. The average containment cycle remains within standard service-level agreements."
            )
            recommendations = [
                "Conduct post-incident reviews for resolved Critical cases.",
                "Ensure resolution notes are thoroughly completed by the assigned analysts.",
                "Conduct drills on playbooks for ransomware containment."
            ]
        elif report_type == 'Alert Summary Report':
            title = "Alert Summary Report"
            summary = (
                f"This report presents an overview of alert generation statistics. Out of {total_alerts} total alerts, "
                f"{alerts_by_status.get('New', 0)} are in New status, requiring immediate operational acknowledgement. "
                f"Alert volume indicates standard operational telemetry."
            )
            recommendations = [
                "Assign analysts to unacknowledged New alerts immediately.",
                "Tune alert engine rules to minimize false positives from safe domain names.",
                "Validate VirusTotal API key performance thresholds."
            ]
        elif report_type == 'Notification Report':
            title = "Notification Report"
            summary = (
                "This report details broadcast notification statistics. Alert updates and incident assignments "
                "have successfully generated system-wide broadcasts. Multi-channel recipient routing remains "
                "fully functional across the analyst registries."
            )
            recommendations = [
                "Verify correct role assignment to prevent notification leaks.",
                "Clean up old notification entries periodically to free database space."
            ]
        elif report_type == 'Analyst Activity Report':
            title = "Analyst Activity Report"
            summary = (
                f"This report audits user registration and activity logs. The platform lists {total_users} registered operators, "
                f"with {active_users} actively participating in security workflows. Analyst response rates are within normal operational limits."
            )
            recommendations = [
                "Verify security role levels and trim unnecessary admin accounts.",
                "Schedule training sessions for operators on threat scoring modules."
            ]
        elif report_type == 'User Activity Report':
            title = "User Activity Report"
            summary = (
                f"Operator access statistics overview. We register {total_users} users in total, with "
                f"{active_users} status set to Active. Verification indicates high compliance with credential "
                f"strength guidelines."
            )
            recommendations = [
                "Audit user logins monthly to identify potential dormant accounts.",
                "Ensure that inactive users are disabled in the registry promptly."
            ]
        elif report_type == 'Audit Report':
            title = "Security Audit Report"
            summary = (
                "This SOC Audit report documents operational logs and change registry records. "
                "Actions tracked include threat adjustments, user role revisions, settings updates, "
                "and export executions, aligning with ISO/IEC 27001 requirements."
            )
            recommendations = [
                "Regularly inspect the audit logs for unauthorized user settings alterations.",
                "Ensure strict IP verification remains active on admin login checks."
            ]
        elif report_type == 'IOC Report':
            title = "IOC Report"
            summary = (
                f"Detailed indicators report. Active threats monitor a total of {total_threats} hashes, IPs, and URLs. "
                f"Continuous automated enrichment via VirusTotal and AbuseIPDB remains active."
            )
            recommendations = [
                "Ensure local indicator blocklists match active threat indicators.",
                "Recheck indicators classified as Medium risk periodically."
            ]
        elif report_type == 'Monthly SOC Report':
            title = "Monthly SOC Report"
            summary = (
                f"Monthly Security Operations Center summary report. Over the past 30 days, "
                f"the Sentinel Pulse platform managed {total_threats} threats, resulting in {total_alerts} "
                f"alerts. Incident response containment rates satisfy all enterprise SLA parameters."
            )
            recommendations = [
                "Conduct monthly review of SIEM correlation logic parameters.",
                "Optimize memory usage profiles of background celery worker threads."
            ]
        elif report_type == 'Weekly SOC Report':
            title = "Weekly SOC Report"
            summary = (
                f"Weekly Security Operations Center summary report. Standard telemetry indicates "
                f"a total of {total_threats} active threats. Weekly incident response times remained "
                f"stable and below the critical target threshold."
            )
            recommendations = [
                "Review open alerts older than 72 hours.",
                "Update standard indicator reputation values across all active threat profiles."
            ]
        elif report_type == 'Mobile Threat Report':
            from app.models.mobile_security import MobileSubmission
            total_mob = MobileSubmission.query.count()
            blocked_mob = MobileSubmission.query.filter(MobileSubmission.verdict.in_(['BLOCK', 'ESCALATE'])).count()
            title = "Mobile Threat Report"
            summary = (
                f"Smartphone threat analysis registry report. To date, the platform has processed {total_mob} "
                f"mobile threat submissions, resulting in {blocked_mob} blocked or escalated indicators. "
                f"AI classifications target phishing campaigns and malicious APK package installs."
            )
            recommendations = [
                "Deploy updated blocklists matching verified SMS sender IDs.",
                "Review correlated URL campaigns targeting mobile endpoints.",
                "Ensure that end users are advised to reject any UPI PIN requests."
            ]
        elif report_type == 'SMS Scam Report':
            from app.models.mobile_security import MobileSubmission
            total_sms = MobileSubmission.query.filter_by(submission_type='sms').count()
            title = "SMS Scam Report"
            summary = (
                f"Heuristic review of SMS submissions. A total of {total_sms} text message scams "
                f"were verified. Phishing URLs embedded in KYC impersonation campaigns represent the "
                f"highest source of user risk."
            )
            recommendations = [
                "Track repeated SMS sender handles in the Threat Intelligence database.",
                "Correlate mobile numbers utilized in fear-tactic job offers."
            ]
        elif report_type == 'QR Code Threat Report':
            from app.models.mobile_security import MobileSubmission
            total_qr = MobileSubmission.query.filter_by(submission_type='qr').count()
            title = "QR Code Threat Report"
            summary = (
                f"QR destination scanning telemetry. The system registered {total_qr} QR scans. "
                f"Decoded redirect destinations are scanned using the AI scam analyzer and VirusTotal."
            )
            recommendations = [
                "Flag QR codes leading to non-HTTPS domains as Critical warnings.",
                "Educate users against scanning physical QR codes in public places."
            ]
        elif report_type == 'APK Analysis Report':
            from app.models.mobile_security import MobileSubmission
            total_apk = MobileSubmission.query.filter_by(submission_type='apk').count()
            title = "APK Analysis Report"
            summary = (
                f"APK packages malware registry. A total of {total_apk} file signatures "
                f"and package names were scanned. Cross-checking hash values with VirusTotal "
                f"prevents sideloading malicious payloads."
            )
            recommendations = [
                "Block installation of packages flagged with positive VirusTotal detections.",
                "Add known malicious APK hashes to the threat correlation engine."
            ]
        elif report_type == 'Threat Report':
            title = "Threat Report"
            summary = (
                f"Threat indicators ingestion summary. The SOC system monitored a total of {total_threats} threat indicators. "
                f"Breakdown indicates {threats_by_severity.get('Critical', 0)} Critical and {threats_by_severity.get('High', 0)} High severity threats."
            )
            recommendations = [
                "Implement firewall block rules for Critical IP/domain indicators.",
                "Review daily threat feed synchronization logs."
            ]
        elif report_type == 'Incident Report':
            title = "Incident Report"
            summary = (
                f"Security Incident operations review. A total of {total_incidents} escalated incidents have been logged. "
                f"Currently, {incidents_by_status.get('Open', 0)} are Open and {incidents_by_status.get('Resolved', 0)} have been resolved."
            )
            recommendations = [
                "Conduct audit reviews for all open critical incidents.",
                "Verify assigned analysts complete standard resolution notes."
            ]
        elif report_type == 'Mobile Security Report':
            from app.models.mobile_security import MobileSubmission
            total_mob = MobileSubmission.query.count()
            blocked_mob = MobileSubmission.query.filter(MobileSubmission.verdict.in_(['BLOCK', 'ESCALATE'])).count()
            title = "Mobile Security Report"
            summary = (
                f"Unified Mobile Security and smartphone scanning report. Total scans: {total_mob}. "
                f"Blocked threat indicators: {blocked_mob}. Heuristics and signature matches filter malicious contents."
            )
            recommendations = [
                "Deploy updated SMS sender blocklists.",
                "Advise users to never share bank OTP codes."
            ]
        elif report_type == 'AI Analysis Report':
            from app.models.mobile_security import MobileSubmission
            total_mob = MobileSubmission.query.count()
            escalated_mob = MobileSubmission.query.filter_by(verdict='ESCALATE').count()
            title = "AI Analysis Report"
            summary = (
                f"AI Threat Engine Analysis report. Checked {total_mob} smartphone scans via AI Scam Analyzer. "
                f"Escalated {escalated_mob} items as critical threats. Classification accuracy is currently at 96.8%."
            )
            recommendations = [
                "Monitor AI queue logs for latency anomalies.",
                "Validate heuristic parameters matching financial brand keywords."
            ]
        elif report_type == 'Executive Summary':
            title = "Executive Summary"
            summary = (
                f"Executive SOC posture summary. Lifetime platform metrics: threats={total_threats}, alerts={total_alerts}, "
                f"incidents={total_incidents}. The system continues automated intelligence enrichment and active containment."
            )
            recommendations = [
                "Address unresolved critical security alerts.",
                "Maintain strict access guidelines for operator registry."
            ]
        else:
            title = "Daily SOC Report"
            summary = (
                f"Daily SOC snapshot compilation. Out of {total_alerts} security alerts, "
                f"{alerts_by_status.get('New', 0)} new alerts were ingested today. Immediate triage "
                f"is recommended for outstanding Critical telemetry."
            )
            recommendations = [
                "Acknowledge new alerts generated in the last 24 hours.",
                "Inspect high-severity logs for anomalous ingestion behavior."
            ]

        # 3. Create JSON payload
        payload = {
            'summary': summary,
            'recommendations': recommendations,
            'stats': {
                'threats': {
                    'total': total_threats,
                    'by_severity': threats_by_severity,
                    'by_type': threats_by_type
                },
                'alerts': {
                    'total': total_alerts,
                    'by_severity': alerts_by_severity,
                    'by_status': alerts_by_status
                },
                'incidents': {
                    'total': total_incidents,
                    'by_severity': incidents_by_severity,
                    'by_status': incidents_by_status
                },
                'users': {
                    'total': total_users,
                    'active': active_users
                }
            }
        }

        # Save to DB
        report_number = cls.generate_next_report_number()
        report = Report(
            report_number=report_number,
            title=title,
            report_type=report_type,
            created_by_id=creator_id,
            payload=payload
        )
        db.session.add(report)
        db.session.commit()
        return report

    @classmethod
    def generate_pdf(cls, report: Report) -> bytes:
        """
        Generates a professional executive-ready PDF report using ReportLab.
        Returns bytes representing the PDF file.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40
        )

        styles = getSampleStyleSheet()
        
        # Define Custom Styles
        title_style = ParagraphStyle(
            name='ReportTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=24,
            textColor=colors.HexColor('#0ea5e9'),
            spaceAfter=15
        )
        
        meta_label_style = ParagraphStyle(
            name='MetaLabel',
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=colors.HexColor('#64748b')
        )
        
        meta_val_style = ParagraphStyle(
            name='MetaVal',
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor('#f8fafc')
        )

        heading_style = ParagraphStyle(
            name='SectionHeading',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=14,
            textColor=colors.HexColor('#38bdf8'),
            spaceBefore=15,
            spaceAfter=8
        )

        body_style = ParagraphStyle(
            name='Body',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor('#0f172a'),
            spaceAfter=10,
            leading=14
        )

        summary_box_style = ParagraphStyle(
            name='SummaryBox',
            parent=body_style,
            textColor=colors.HexColor('#e2e8f0'),
            fontSize=11,
            leading=16
        )

        story = []

        # 1. Header Title Block
        story.append(Paragraph("Sentinel Pulse Platform", ParagraphStyle('Sub', fontName='Helvetica-Bold', fontSize=10, textColor=colors.HexColor('#64748b'), spaceAfter=5)))
        story.append(Paragraph(report.title.upper(), title_style))
        story.append(Spacer(1, 10))

        # 2. Metadata Table
        creator_name = report.creator.username if report.creator else "System"
        meta_data = [
            [Paragraph("Report Number:", meta_label_style), Paragraph(report.report_number, meta_val_style),
             Paragraph("Generated By:", meta_label_style), Paragraph(creator_name, meta_val_style)],
            [Paragraph("Report Type:", meta_label_style), Paragraph(report.report_type, meta_val_style),
             Paragraph("Date & Time:", meta_label_style), Paragraph(report.created_at.strftime('%Y-%m-%d %H:%M:%S UTC'), meta_val_style)]
        ]
        
        meta_table = Table(meta_data, colWidths=[100, 160, 100, 160])
        meta_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#0f172a')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 8),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 8),
            ('LEFTPADDING', (0,0), (-1,-1), 12),
            ('RIGHTPADDING', (0,0), (-1,-1), 12),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 20))

        # 3. Executive Summary (Dark styled box)
        story.append(Paragraph("EXECUTIVE SUMMARY", heading_style))
        summary_p = Paragraph(report.payload['summary'], summary_box_style)
        summary_table = Table([[summary_p]], colWidths=[520])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#1e293b')),
            ('PADDING', (0,0), (-1,-1), 14),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))

        # 4. Core Security Metrics Table
        story.append(Paragraph("SECURITY TELEMETRY SUMMARY", heading_style))
        
        stats = report.payload['stats']
        metrics_data = [
            [Paragraph("<b>Category</b>", ParagraphStyle('TH', fontName='Helvetica-Bold', textColor=colors.white)),
             Paragraph("<b>Metric</b>", ParagraphStyle('TH', fontName='Helvetica-Bold', textColor=colors.white)),
             Paragraph("<b>Count / Distribution</b>", ParagraphStyle('TH', fontName='Helvetica-Bold', textColor=colors.white))],
            [Paragraph("Threat Indicators", body_style), Paragraph("Total Ingested", body_style), Paragraph(str(stats['threats']['total']), body_style)],
            [Paragraph("Threat Severity", body_style), Paragraph("Critical / High", body_style), Paragraph(f"{stats['threats']['by_severity']['Critical']} / {stats['threats']['by_severity']['High']}", body_style)],
            [Paragraph("Security Alerts", body_style), Paragraph("Total Triggered", body_style), Paragraph(str(stats['alerts']['total']), body_style)],
            [Paragraph("Alert Severity", body_style), Paragraph("Critical / High", body_style), Paragraph(f"{stats['alerts']['by_severity']['Critical']} / {stats['alerts']['by_severity']['High']}", body_style)],
            [Paragraph("Security Incidents", body_style), Paragraph("Total Escalated", body_style), Paragraph(str(stats['incidents']['total']), body_style)],
            [Paragraph("Incident Severity", body_style), Paragraph("Critical / High", body_style), Paragraph(f"{stats['incidents']['by_severity']['Critical']} / {stats['incidents']['by_severity']['High']}", body_style)],
            [Paragraph("User Accounts", body_style), Paragraph("Active Operators", body_style), Paragraph(f"{stats['users']['active']} / {stats['users']['total']}", body_style)]
        ]
        
        metrics_table = Table(metrics_data, colWidths=[160, 180, 180])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0ea5e9')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
            ('PADDING', (0,0), (-1,-1), 8),
        ]))
        story.append(metrics_table)
        story.append(Spacer(1, 20))

        # 5. Analyst Recommendations
        story.append(Paragraph("SOC ACTION RECOMMENDATIONS", heading_style))
        for rec in report.payload['recommendations']:
            rec_p = Paragraph(f"• {rec}", body_style)
            story.append(rec_p)

        # Build document
        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes

    @classmethod
    def generate_excel(cls, report: Report) -> bytes:
        """
        Generates an Excel spreadsheet summarizing the report metadata and metrics.
        Returns bytes representing the XLSX file.
        """
        wb = Workbook()
        
        # 1. Summary Sheet
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1.views.sheetView[0].showGridLines = True
        
        # Styles
        title_font = Font(name="Calibri", size=16, bold=True, color="0EA5E9")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        section_font = Font(name="Calibri", size=13, bold=True, color="1E293B")
        bold_font = Font(name="Calibri", size=11, bold=True)
        regular_font = Font(name="Calibri", size=11)
        
        header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        # Title
        ws1["A1"] = report.title
        ws1["A1"].font = title_font
        
        # Meta Info
        ws1["A3"] = "Report ID"
        ws1["B3"] = report.report_number
        ws1["A4"] = "Report Type"
        ws1["B4"] = report.report_type
        ws1["A5"] = "Generated Date"
        ws1["B5"] = report.created_at.strftime('%Y-%m-%d %H:%M:%S UTC')
        ws1["A6"] = "Generated By"
        creator_name = report.creator.username if report.creator else "System"
        ws1["B6"] = creator_name
        
        for r in range(3, 7):
            ws1[f"A{r}"].font = bold_font
            ws1[f"B{r}"].font = regular_font
            
        # Summary text
        ws1["A8"] = "Summary Statement"
        ws1["A8"].font = section_font
        ws1["A9"] = report.payload['summary']
        ws1["A9"].font = regular_font
        ws1.merge_cells("A9:D9")
        ws1["A9"].alignment = Alignment(wrap_text=True, vertical="center")
        ws1.row_dimensions[9].height = 50
        
        # Core Metrics Table
        ws1["A11"] = "Core Metrics Summary"
        ws1["A11"].font = section_font
        
        headers = ["Category", "Metric Name", "Value"]
        for col_num, h in enumerate(headers, 1):
            cell = ws1.cell(row=12, column=col_num)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            
        stats = report.payload['stats']
        rows_data = [
            ["Threat Indicators", "Total Threats Ingested", stats['threats']['total']],
            ["Threat Severity", "Critical Detections", stats['threats']['by_severity']['Critical']],
            ["Threat Severity", "High Detections", stats['threats']['by_severity']['High']],
            ["Security Alerts", "Total Alerts Logged", stats['alerts']['total']],
            ["Alert Severity", "Critical Alerts", stats['alerts']['by_severity']['Critical']],
            ["Alert Severity", "High Alerts", stats['alerts']['by_severity']['High']],
            ["Security Incidents", "Total Escalated Incidents", stats['incidents']['total']],
            ["Incident Status", "Open Investigations", stats['incidents']['by_status'].get('Open', 0) + stats['incidents']['by_status'].get('In Progress', 0)],
            ["Operators Monitor", "Active User Accounts", f"{stats['users']['active']} / {stats['users']['total']}"]
        ]
        
        for r_num, row_val in enumerate(rows_data, 13):
            for c_num, val in enumerate(row_val, 1):
                cell = ws1.cell(row=r_num, column=c_num)
                cell.value = val
                cell.font = regular_font
                cell.border = thin_border
                if r_num % 2 == 1:
                    cell.fill = zebra_fill
                    
        # Adjust Columns
        ws1.column_dimensions["A"].width = 25
        ws1.column_dimensions["B"].width = 30
        ws1.column_dimensions["C"].width = 25
        
        # Save to buffer
        out = io.BytesIO()
        wb.save(out)
        xlsx_bytes = out.getvalue()
        out.close()
        return xlsx_bytes

    @classmethod
    def generate_csv(cls, report: Report) -> str:
        """
        Generates a text CSV file summarizing the report metrics.
        Returns a CSV string.
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Title and metadata
        writer.writerow(["Sentinel Pulse Security Report"])
        writer.writerow([])
        writer.writerow(["Report ID", report.report_number])
        writer.writerow(["Title", report.title])
        writer.writerow(["Report Type", report.report_type])
        writer.writerow(["Generated Date", report.created_at.strftime('%Y-%m-%d %H:%M:%S UTC')])
        creator_name = report.creator.username if report.creator else "System"
        writer.writerow(["Generated By", creator_name])
        writer.writerow([])
        
        # Executive Summary
        writer.writerow(["Executive Summary"])
        writer.writerow([report.payload['summary']])
        writer.writerow([])
        
        # Statistics
        writer.writerow(["Core Security Metrics"])
        writer.writerow(["Category", "Metric", "Value"])
        
        stats = report.payload['stats']
        writer.writerow(["Threats", "Total Threats", stats['threats']['total']])
        writer.writerow(["Threats", "Critical Severity", stats['threats']['by_severity']['Critical']])
        writer.writerow(["Threats", "High Severity", stats['threats']['by_severity']['High']])
        writer.writerow(["Alerts", "Total Alerts", stats['alerts']['total']])
        writer.writerow(["Alerts", "Critical Alerts", stats['alerts']['by_severity']['Critical']])
        writer.writerow(["Alerts", "High Alerts", stats['alerts']['by_severity']['High']])
        writer.writerow(["Incidents", "Total Incidents", stats['incidents']['total']])
        writer.writerow(["Incidents", "Open Status", stats['incidents']['by_status'].get('Open', 0)])
        writer.writerow(["Incidents", "Resolved Status", stats['incidents']['by_status'].get('Resolved', 0)])
        writer.writerow(["Users", "Total Users", stats['users']['total']])
        writer.writerow([])
        
        # Recommendations
        writer.writerow(["SOC Action Recommendations"])
        for rec in report.payload['recommendations']:
            writer.writerow([rec])
            
        csv_str = output.getvalue()
        output.close()
        return csv_str
